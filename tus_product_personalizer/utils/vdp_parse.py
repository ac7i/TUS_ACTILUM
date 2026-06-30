"""Parse CSV / XLSX files for Variable Data Printing."""

import csv
import io
import logging

_logger = logging.getLogger(__name__)


def parse_vdp_spreadsheet(file_bytes, filename=""):
    """Return ``{'fields': [...], 'records': [{field: value, ...}, ...]}``.

  Supports UTF-8 CSV and XLSX (first sheet, first row = headers).
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(file_bytes)
    return _parse_csv(file_bytes)


def _normalize_header(header):
    return (header or "").strip()


def _parse_csv(file_bytes):
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return {"fields": [], "records": []}
    
    # Map raw header -> normalized header, filtering out empty ones
    header_mapping = {}
    fields = []
    for h in reader.fieldnames:
        norm_h = _normalize_header(h)
        if norm_h:
            header_mapping[h] = norm_h
            if norm_h not in fields:
                fields.append(norm_h)

    records = []
    for row in reader:
        record = {}
        for raw_h, norm_h in header_mapping.items():
            record[norm_h] = (row.get(raw_h) or "").strip()
        if any(record.values()):
            records.append(record)
    return {"fields": fields, "records": records}


def _parse_xlsx(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        _logger.error("openpyxl is required for VDP XLSX import: %s", exc)
        return {"fields": [], "records": [], "error": "openpyxl not installed"}

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"fields": [], "records": []}

    # Map from column index to normalized header name
    column_mapping = {}
    fields = []
    for idx, cell in enumerate(rows[0]):
        norm_h = _normalize_header(str(cell)) if cell is not None else ""
        if norm_h:
            column_mapping[idx] = norm_h
            if norm_h not in fields:
                fields.append(norm_h)

    if not fields:
        return {"fields": [], "records": []}

    records = []
    for row in rows[1:]:
        record = {}
        for idx, field in column_mapping.items():
            val = row[idx] if idx < len(row) else None
            record[field] = "" if val is None else str(val).strip()
        if any(record.values()):
            records.append(record)
    return {"fields": fields, "records": records}
