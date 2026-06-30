# -*- coding: utf-8 -*-
"""Parse design-template bulk import spreadsheets (XLSX)."""
import base64
import io
import zipfile
from io import BytesIO

try:
    from openpyxl import load_workbook, Workbook
except ImportError:  # pragma: no cover - guarded in wizard
    load_workbook = None
    Workbook = None

HEADER_ALIASES = {
    "name": {"name", "template_name", "title", "template"},
    "sequence": {"sequence", "seq", "order", "sort"},
    "design_format": {"design_format", "format", "type", "file_type"},
    "design_file": {
        "design_file",
        "design_path",
        "svg_file",
        "file",
        "design",
        "source_file",
    },
    "preview_file": {"preview_file", "preview_path", "preview", "thumbnail", "image"},
    "is_default": {"is_default", "default", "default_template"},
    "active": {"active", "enabled", "publish"},
}


def _norm_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def _map_headers(header_row):
    mapping = {}
    for col_idx, cell in enumerate(header_row):
        key = _norm_header(cell)
        if not key:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if key in aliases:
                mapping[field] = col_idx
                break
    return mapping


def _cell(row, col_idx):
    if col_idx is None or col_idx >= len(row):
        return None
    value = row[col_idx]
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def _parse_bool(value, default=False):
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(int(value))
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _parse_int(value, default=10):
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def parse_design_template_xlsx(xlsx_bytes):
    """Return list of row dicts from the first worksheet."""
    if not load_workbook:
        raise ImportError("openpyxl")
    wb = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header_map = _map_headers(rows[0])
    if "name" not in header_map or "design_file" not in header_map:
        raise ValueError(
            "Missing required columns. The sheet must include at least: "
            "name, design_file"
        )
    parsed = []
    for row in rows[1:]:
        if not row or not any(cell not in (None, "") for cell in row):
            continue
        name = _cell(row, header_map.get("name"))
        design_file = _cell(row, header_map.get("design_file"))
        if not name and not design_file:
            continue
        parsed.append({
            "name": name,
            "sequence": _parse_int(_cell(row, header_map.get("sequence")), 10),
            "design_format": _cell(row, header_map.get("design_format")),
            "design_file": design_file,
            "preview_file": _cell(row, header_map.get("preview_file")),
            "is_default": _parse_bool(_cell(row, header_map.get("is_default")), False),
            "active": _parse_bool(_cell(row, header_map.get("active")), True),
        })
    return parsed


def build_zip_file_index(zip_bytes):
    """Map normalized relative paths inside a ZIP to raw file bytes."""
    index = {}
    with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            if name.endswith("/"):
                continue
            data = zf.read(name)
            norm = name.replace("\\", "/").lstrip("./")
            index[norm] = data
            index[norm.split("/")[-1]] = data
    return index


def read_zip_member(zip_index, path):
    if not path:
        return None
    norm = str(path).replace("\\", "/").lstrip("./")
    if norm in zip_index:
        return zip_index[norm]
    basename = norm.split("/")[-1]
    return zip_index.get(basename)


def infer_design_format(filename, explicit_format=None):
    fmt = (explicit_format or "").strip().lower()
    if fmt in {"svg", "fabric_json", "json"}:
        return "fabric_json" if fmt in {"fabric_json", "json"} else "svg"
    lower = (filename or "").lower()
    if lower.endswith(".json"):
        return "fabric_json"
    return "svg"


def generate_sample_xlsx_bytes():
    if not Workbook:
        raise ImportError("openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "Templates"
    ws.append([
        "name",
        "sequence",
        "design_format",
        "design_file",
        "preview_file",
        "is_default",
        "active",
    ])
    ws.append([
        "Classic Blue Card",
        10,
        "svg",
        "designs/classic-blue.svg",
        "previews/classic-blue.png",
        "no",
        "yes",
    ])
    ws.append([
        "Modern Minimal",
        20,
        "svg",
        "designs/modern-minimal.svg",
        "previews/modern-minimal.png",
        "no",
        "yes",
    ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
